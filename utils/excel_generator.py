import os
import openpyxl
from openpyxl.utils import get_column_letter

def generate_excel_from_json(json_data, output_filepath):
    """
    Generates an Excel file from the structured JSON data.
    This is a basic implementation; more complex structures might need deeper recursion.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Chapter Content"

    headers = ["Type", "Name/Title", "Content/Description", "Details (e.g., Steps, Rows, Caption)"]
    sheet.append(headers)

    row_num = 2 

    if "content" not in json_data or not isinstance(json_data["content"], list):
        print("Invalid JSON format: 'content' key missing or not a list.")
        return False

    chapter_title = json_data.get("chapter_title", "Unknown Chapter")
    sheet[f'A{row_num}'] = "Chapter Title"
    sheet[f'B{row_num}'] = chapter_title
    row_num += 1

    def write_elements_to_sheet(elements, level=0):
        nonlocal row_num 
        indent = "  " * level 

        for element in elements:
            if element.get("type") in ["topic", "sub_topic"]:
                sheet.cell(row=row_num, column=1, value=f"{indent}{element['type'].capitalize()}")
                sheet.cell(row=row_num, column=2, value=element.get("name", "N/A"))
                row_num += 1
                if "elements" in element and isinstance(element["elements"], list):
                    write_elements_to_sheet(element["elements"], level + 1)
            else:
                sheet.cell(row=row_num, column=1, value=f"{indent}{element.get('type', 'Unknown').capitalize()}")
                sheet.cell(row=row_num, column=2, value=element.get("name", ""))
                if element.get("type") == "paragraph":
                    sheet.cell(row=row_num, column=3, value=element.get("text", ""))
                elif element.get("type") in ["image", "diagram"]:
                    sheet.cell(row=row_num, column=3, value=element.get("description", ""))
                    sheet.cell(row=row_num, column=4, value=element.get("caption", ""))
                elif element.get("type") == "table":
                    sheet.cell(row=row_num, column=3, value=element.get("caption", ""))
                    rows_content = ""
                    if "rows" in element and isinstance(element["rows"], list):
                        for row in element["rows"]:
                            rows_content += "| " + " | ".join(map(str, row)) + " |\n"
                    sheet.cell(row=row_num, column=4, value=rows_content.strip())
                elif element.get("type") == "activity":
                    sheet.cell(row=row_num, column=3, value=element.get("description", ""))
                    steps_content = "\n".join(element.get("steps", []))
                    sheet.cell(row=row_num, column=4, value=steps_content)
                elif element.get("type") == "question":
                    sheet.cell(row=row_num, column=3, value=element.get("text", ""))
                elif element.get("type") == "example":
                     sheet.cell(row=row_num, column=3, value=element.get("text", ""))
                elif element.get("type") == "external_source":
                    sheet.cell(row=row_num, column=3, value=element.get("text", ""))


                row_num += 1

    write_elements_to_sheet(json_data["content"])

   
    for col in sheet.columns:
        max_length = 0
        column = col[0].column #
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    try:
        workbook.save(output_filepath)
        print(f"Excel file saved to: {output_filepath}")
        return True
    except Exception as e:
        print(f"Error saving Excel file {output_filepath}: {e}")
        return False

if __name__ == '__main__':
    
    dummy_json_data = {
        "chapter_title": "Sample Chapter on Photosynthesis",
        "content": [
            {
                "type": "topic",
                "name": "What is Photosynthesis?",
                "elements": [
                    {
                        "type": "paragraph",
                        "text": "Photosynthesis is the process by which green plants and some other organisms use sunlight to synthesize foods with chlorophyll. This process uses carbon dioxide and water as reactants."
                    },
                    {
                        "type": "image",
                        "description": "Diagram of a plant leaf showing stomata.",
                        "caption": "Fig. 1.1: Cross-section of a leaf."
                    },
                    {
                        "type": "sub_topic",
                        "name": "Requirements for Photosynthesis",
                        "elements": [
                            {
                                "type": "paragraph",
                                "text": "For photosynthesis to occur, plants need sunlight, chlorophyll, water, and carbon dioxide."
                            },
                            {
                                "type": "table",
                                "caption": "Table 1.1: Key Components of Photosynthesis",
                                "rows": [
                                    ["Component", "Source"],
                                    ["Sunlight", "Sun"],
                                    ["Water", "Soil"],
                                    ["Carbon Dioxide", "Air"]
                                ]
                            }
                        ]
                    },
                    {
                        "type": "activity",
                        "description": "To demonstrate that sunlight is essential for photosynthesis.",
                        "steps": [
                            "Take a potted plant and keep it in the dark for three days.",
                            "Cover one of its leaves with black paper.",
                            "Keep the plant in sunlight for a few hours.",
                            "Perform starch test on both leaves."
                        ]
                    },
                    {
                        "type": "question",
                        "text": "Why is chlorophyll important for plants?"
                    }
                ]
            }
        ]
    }

    output_dir = "../data/output_excel/"
    os.makedirs(output_dir, exist_ok=True)
    output_excel_path = os.path.join(output_dir, "sample_chapter_extract.xlsx")
    generate_excel_from_json(dummy_json_data, output_excel_path)